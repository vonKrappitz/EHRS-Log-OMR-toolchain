# -*- coding: utf-8 -*-
# SPDX-License-Identifier: Apache-2.0
# Copyright 2026 Maciej Kasperek
"""
EHRS Log OMR reader (proof of concept).

Reads scanned EHRS Log forms, finds the four solid corner fiducials, maps the
layout coordinate frame onto the scan via a homography, samples each bubble
interior, and writes the recovered records to an Excel workbook (one sheet per
stream: pilot / doctor / dispatcher).

Anonymity: doctor cards are pooled and SHUFFLED before sequential serial
numbers (D01) are assigned, so the serial carries no order, date, or link.

Usage:
    python omr_reader.py --layout EHRS_forms_layout_full.json \\
        --root scans/ --out EHRS_capture.xlsx
where scans/ holds subfolders pilot/ doctor/ dispatcher/ with page images
(PNG/JPG). Auto fields and the doctor serial are left for the back end / scan
step; everything bubbled is recovered here.
"""
import argparse
import glob
import json
import os
import random

import cv2
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

FILL_THRESH = 0.35      # interior dark fraction above which a bubble counts as marked
INTERIOR = 0.62         # sample this fraction of the radius (skip the printed ring)


# ----------------------------- layout helpers -----------------------------
def load_layout(path):
    with open(path) as f:
        return json.load(f)


def fiducials_ordered(form_layout):
    """Layout fiducial centres (x, y_top in pt) ordered TL, TR, BL, BR."""
    pts = form_layout["fiducials"]
    s = sorted(pts, key=lambda p: p[1])
    top = sorted(s[:2], key=lambda p: p[0])
    bot = sorted(s[2:], key=lambda p: p[0])
    return [top[0], top[1], bot[0], bot[1]]


# ----------------------------- image pipeline -----------------------------
def load_gray(path):
    img = cv2.imread(path, cv2.IMREAD_GRAYSCALE)
    if img is None:
        raise RuntimeError("cannot read image: %s" % path)
    return img


def find_fiducials(gray):
    """Return the four corner fiducial centres in pixels, ordered TL, TR, BL, BR."""
    h, w = gray.shape
    _, bw = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    bw = cv2.morphologyEx(bw, cv2.MORPH_CLOSE, np.ones((3, 3), np.uint8))
    cnts, _ = cv2.findContours(bw, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    area = w * h
    cand = []
    for c in cnts:
        a = cv2.contourArea(c)
        if a < area * 0.0003 or a > area * 0.01:
            continue
        x, y, ww, hh = cv2.boundingRect(c)
        ar = ww / float(hh) if hh else 0
        extent = a / float(ww * hh) if ww * hh else 0
        if 0.6 < ar < 1.6 and extent > 0.7:   # solid square-ish blob
            cand.append((x + ww / 2.0, y + hh / 2.0, a))
    if len(cand) < 4:
        raise RuntimeError("found %d fiducial candidates, need 4" % len(cand))
    corners = [(0, 0), (w, 0), (0, h), (w, h)]
    chosen, used = [], set()
    for (rx, ry) in corners:
        best, bestd, bi = None, 1e18, -1
        for i, (cx, cy, a) in enumerate(cand):
            if i in used:
                continue
            d = (cx - rx) ** 2 + (cy - ry) ** 2
            if d < bestd:
                bestd, best, bi = d, (cx, cy), i
        used.add(bi)
        chosen.append(best)
    return np.array(chosen, dtype=np.float32)


def build_homography(form_layout, fid_px):
    src = np.array(fiducials_ordered(form_layout), dtype=np.float32)
    H, _ = cv2.findHomography(src, fid_px)
    return H


def map_pt(H, x, y):
    v = H @ np.array([x, y, 1.0])
    return v[0] / v[2], v[1] / v[2]


def fill_ratio(gray, H, cx, cy, r):
    px, py = map_pt(H, cx, cy)
    rx, ry = map_pt(H, cx + r, cy)
    rad = max(3.0, ((rx - px) ** 2 + (ry - py) ** 2) ** 0.5) * INTERIOR
    h, w = gray.shape
    x0, x1 = int(px - rad), int(px + rad)
    y0, y1 = int(py - rad), int(py + rad)
    x0, y0 = max(0, x0), max(0, y0)
    x1, y1 = min(w - 1, x1), min(h - 1, y1)
    patch = gray[y0:y1, x0:x1]
    if patch.size == 0:
        return 0.0
    return float((patch < 128).mean())


# ----------------------------- form reading -----------------------------
def read_form(gray, form_layout, H, fill_thresh=FILL_THRESH):
    out, flags = {}, {}
    for fid, spec in form_layout["fields"].items():
        t = spec["type"]
        if t in ("auto", "qr_serial"):
            out[fid] = ""                       # back end / scan step fills these
        elif t == "digit_key":
            digits, multi = "", False
            for place in spec["places"]:
                ratios = [(b["digit"], fill_ratio(gray, H, b["cx"], b["cy"], b["r"]))
                          for b in place["bubbles"]]
                marked = [d for d, rr in ratios if rr >= fill_thresh]
                if len(marked) == 1:
                    digits += str(marked[0])
                elif not marked:
                    digits += "0"               # blank leading/internal row = 0
                else:
                    digits += str(max(ratios, key=lambda z: z[1])[0])
                    multi = True
            out[fid] = digits
            if multi:
                flags[fid] = "multi"
        elif t == "bubble_single":
            ratios = [(o["value"], fill_ratio(gray, H, o["cx"], o["cy"], o["r"]))
                      for o in spec["options"]]
            marked = [v for v, rr in ratios if rr >= fill_thresh]
            if len(marked) == 1:
                out[fid] = marked[0]
            elif not marked:
                out[fid] = ""
                flags[fid] = "blank"
            else:
                out[fid] = max(ratios, key=lambda z: z[1])[0]
                flags[fid] = "multi"
        elif t == "digit_count":
            counts = {}
            for sub, sd in spec["subcounts"].items():
                digits, ok = "", True
                for place in sd["places"]:
                    ratios = [(b["digit"], fill_ratio(gray, H, b["cx"], b["cy"], b["r"]))
                              for b in place["bubbles"]]
                    marked = [d for d, rr in ratios if rr >= fill_thresh]
                    if len(marked) == 1:
                        digits += str(marked[0])
                    elif not marked:
                        digits += "0"
                    else:
                        digits += str(max(ratios, key=lambda z: z[1])[0])
                        ok = False
                counts[sub] = int(digits) if digits else 0
                if not ok:
                    flags["%s.%s" % (fid, sub)] = "multi"
            out[fid] = counts
    return out, flags


def process_images(layout, form_key, image_paths, fill_thresh=FILL_THRESH):
    fl = layout["forms"][form_key]
    records = []
    for p in sorted(image_paths):
        gray = load_gray(p)
        fid_px = find_fiducials(gray)
        H = build_homography(fl, fid_px)
        rec, flags = read_form(gray, fl, H, fill_thresh)
        rec["_source"] = os.path.basename(p)
        rec["_flags"] = ";".join("%s:%s" % (k, v) for k, v in flags.items())
        records.append(rec)
    if form_key == "doctor":
        random.shuffle(records)                 # pool shuffle BEFORE serial assignment
        for i, rec in enumerate(records, 1):
            rec["D01"] = "EHRS-%05d" % i
    return records


# ----------------------------- excel output -----------------------------
def columns_for(form_layout, form_key=None):
    cols = []
    for fid, spec in form_layout["fields"].items():
        if spec["type"] == "digit_count":
            for sub in spec["subcounts"]:
                cols.append("%s_%s" % (fid, sub))
        else:
            cols.append(fid)
    cols.append("_flags")
    if form_key != "doctor":          # doctor card stays unlinked: no scan-file trace
        cols.append("_source")
    return cols


def write_excel(all_records, layout, out_path):
    wb = Workbook()
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="2A3A73")
    first = True
    for form_key in ("pilot", "doctor", "dispatcher"):
        if form_key not in layout["forms"]:
            continue
        fl = layout["forms"][form_key]
        cols = columns_for(fl, form_key)
        ws = wb.active if first else wb.create_sheet()
        ws.title = form_key
        first = False
        for j, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=j, value=col)
            cell.font = head_font
            cell.fill = head_fill
        for i, rec in enumerate(all_records.get(form_key, []), 2):
            for j, col in enumerate(cols, 1):
                if "_" in col and col.split("_")[0] in fl["fields"] \
                        and fl["fields"][col.split("_")[0]]["type"] == "digit_count":
                    fid, sub = col.split("_", 1)
                    val = rec.get(fid, {}).get(sub, "")
                else:
                    val = rec.get(col, "")
                ws.cell(row=i, column=j, value=val)
        ws.freeze_panes = "A2"
        for j, col in enumerate(cols, 1):
            ws.column_dimensions[ws.cell(row=1, column=j).column_letter].width = max(10, len(col) + 2)
    wb.save(out_path)


# ----------------------------- cli -----------------------------
def main():
    ap = argparse.ArgumentParser(description="EHRS Log OMR reader (PoC)")
    ap.add_argument("--layout", required=True, help="layout JSON (e.g. EHRS_forms_layout_full.json)")
    ap.add_argument("--root", help="folder with pilot/ doctor/ dispatcher/ subfolders of images")
    ap.add_argument("--form", choices=["pilot", "doctor", "dispatcher"], help="single form type")
    ap.add_argument("--images", help="glob of images for --form mode")
    ap.add_argument("--out", default="EHRS_capture.xlsx")
    ap.add_argument("--seed", type=int, default=None, help="shuffle seed (reproducibility)")
    args = ap.parse_args()
    if args.seed is not None:
        random.seed(args.seed)
    layout = load_layout(args.layout)
    all_records = {}
    if args.root:
        for fk in ("pilot", "doctor", "dispatcher"):
            d = os.path.join(args.root, fk)
            imgs = []
            for ext in ("*.png", "*.jpg", "*.jpeg", "*.tif", "*.tiff"):
                imgs += glob.glob(os.path.join(d, ext))
            if imgs:
                all_records[fk] = process_images(layout, fk, imgs)
    elif args.form and args.images:
        all_records[args.form] = process_images(layout, args.form, glob.glob(args.images))
    else:
        ap.error("provide --root, or --form together with --images")
    write_excel(all_records, layout, args.out)
    n = sum(len(v) for v in all_records.values())
    print("wrote %s  (%d records across %d streams)" % (args.out, n, len(all_records)))


if __name__ == "__main__":
    main()
